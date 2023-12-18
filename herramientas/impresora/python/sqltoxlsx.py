#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
impresora a xlsx:
uso:
	sqltoxlsx.py <jsonfile>
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from datetime import datetime
import requests
import json
import sys
from StringIO import StringIO

jsonfilename  = sys.argv[1]

with open(jsonfilename) as json_file:
    js = json.load(json_file)

wb = Workbook()

ft = Font(color="FFFFFF")

n=0
for key in js["datasets"]:
	fields = js["datasets"][key]["fields"]
	lf = len(fields)

	data = js["datasets"][key]["data"]
	ld = len(data)

	if (ld==0):
		continue

	if n==0:
		ws = wb.active
		ws.title = key
	else:
		ws = wb.create_sheet(key)
	n=n+1

	
	for i in range(0,lf):
		ws.cell(row=1,column=i+1,value=fields[i])

	for i in range(0,ld):
		lc = len(data[0])
		for j in range(0,lc):
			tp = js["datasets"][key]["types"][j]
			
			if tp == "N":
				try:
					ws.cell(row=2+i,column=j+1).value=float(data[i][j])
				except:
					ws.cell(row=2+i,column=j+1).value=0
			elif tp == "D":
				v = data[i][j]
				try:
					if len(v) == 10:
						ws.cell(row=2+i,column=j+1).value=datetime.strptime(data[i][j], '%Y-%m-%d')
						ws.cell(row=2+i,column=j+1).number_format='dd/mm/yyyy'
					elif len(v) > 19:	
						ws.cell(row=2+i,column=j+1).value=datetime.strptime(data[i][j], '%Y-%m-%d %H:%M:%S')
						ws.cell(row=2+i,column=j+1).number_format='dd/mm/yyyy hh:mm:ss'
					else:
						ws.cell(row=2+i,column=j+1).value=data[i][j]
				except:
					ws.cell(row=2+i,column=j+1).value=data[i][j]
			else:
				ws.cell(row=2+i,column=j+1).value=data[i][j]

	ws.auto_filter.ref = ws.dimensions
	

	col_range = ws.max_column
	for col in range(1, col_range + 1):
		cell_header = ws.cell(1, col)
		cell_header.font = ft
		cell_header.fill = PatternFill(start_color="366092", end_color="366092", fill_type = "solid")

	ws.freeze_panes = ws['A2']

output = StringIO()
wb.save(output)

sys.stdout.write(output.getvalue())
